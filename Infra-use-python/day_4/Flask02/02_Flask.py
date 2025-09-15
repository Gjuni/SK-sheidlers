from flask import Flask
from flask import render_template ## 포트포워딩 시킴. 해당 templates/'file_name.html' 로 이동 시킨다.
from flask import redirect
from flask import request
from flask import send_file
import os
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
import time

app = Flask(__name__)


@app.route("/")
def index():
    return render_template('index.html')

@app.route("/upload", methods=['GET', 'POST'])
def upload(): # index.html에서 라우팅 되는 부분 유의 해야함.
    file = request.files["file"] # file을 받을 때 request.files로 받게 됨. (아까는 form이었음.)

    ## 바로 처리하려면 base4로 처리해아한다.
    ## 사용자가 파일 올렸던 것을 upload라는 폴더에 저장헀다가 가져와서 번역하는 것이 안전함.
    ## 이 방법이 아니면 메모리에 저장해서 접근 한 후 다시 번역하는 과정이 필요.
    file.save(os.path.join("uploads", file.filename))

    wb = load_workbook(os.path.join("uploads", file.filename))
    ws = wb.active

    for row in ws.iter_rows() : ## iter_rows 데이터의 갯수가 몇개인지 모를 경우 행 기준으로 없으면 끊기게 만드는 것. 열은 띄엄띄엄 데이터가 존재할 수 있기 때문
        # time.sleep(0.1)
        for cell in row:
            translate = GoogleTranslator(source='ko', target='en').translate(cell.value) ## .translate(value) cell의 값을 번역해라.
            cell.value = translate ## 번역 했던 값을 바꾼다.

    wb.save("result_en.xlsx")
    return render_template('result.html', wb=wb)

@app.route('/download')
def download_report():
    return send_file('result_en.xlsx', as_attachment=True)

if __name__ == '__main__':  # app 실행
    app.run(debug=True)     # 개발자 모드로 실행. 외부 실행 X localhost로만 실행
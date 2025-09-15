import feedparser
from openpyxl import Workbook

url = "https://www.dailysecu.com/rss/allArticle.xml"

feed = feedparser.parse(url)

titles = []
links = []
descriptions = []
authors = []
pubDates = []

wb = Workbook()
ws = wb.active

for entry in feed.entries:
    titles.append(entry.title) ## 각 Tag 값 별로 값을 가져옴
    links.append(entry.link)
    descriptions.append(entry.description)
    authors.append(entry.author)
    pubDates.append(entry.published)

for i in range(len(titles)) :
    ws.cell(row=i+2, column=1).value = i+1
    ws.cell(row=i+2, column=2).value = titles[i]
    ws.cell(row=i+2, column=3).value = links[i]
    ws.cell(row=i+2, column=4).value = descriptions[i]
    ws.cell(row=i+2, column=5).value = authors[i]
    ws.cell(row=i+2, column=6).value = pubDates[i]

wb.save('test.xlsx')

print("파일이 생성")
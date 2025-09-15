import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url = "https://www.malware-traffic-analysis.net/2023/index.html"

header_info = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36'}
r = requests.get(url, headers=header_info)
soup = BeautifulSoup(r.text, 'html.parser')

tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

#워크북
wb = Workbook()
ws = wb.active

ws['A1'] = "설명"
ws['B1'] = "URL 링크"


for tag in tags:
    ws.append([tag.text, f"https://www.malware-traffic-analysis.net/2023/{tag.get('href')}"])

wb.save("malwares01.xlsx")
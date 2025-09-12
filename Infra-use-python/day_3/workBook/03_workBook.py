## 엑셀에서 데이터 불러오기

from openpyxl import load_workbook

wb = load_workbook("excel_data2.xlsx", data_only=True) ## Data값만 가져옴. 시간 문제 해결 할 수 있음
ws = wb.active # 데이터를 불러오게 된 wb를 활성화 하게 만듬
## 시트 별로 데이터를 불러옴

cell = ws["A1:E7"]

for row in cell: # 행 단위 row
    result = [] ## 저장
    # values = [cell.value(for column in row)]

    for column in row: # 열 단위 column
            result.append(column.value) # cell value값을 가져와서 list형식에 저장함
    print(result)

## 시간 처리가 문제가 된다. (총 시간 '=B2*D2')
# ['학습 과정', '시간', '교육방식', '수강 인원', '총 시간']
# ['인프라 활용을 위한 파이썬', 56, '온라인', 50, '=B2*D2']
# ['애플리케이션보안', 56, '온라인', 50, '=B3*D3']
# ['시스템/네트워크 보안 기술', 56, '온라인', 50, '=B4*D4']
# ['클라우드 보안 기술', 56, '온라인', 50, '=B5*D5']
# ['데이터 3법과 개인정보보호', 32, '온라인', 50, '=B6*D6']
# ['모듈 프로젝트', 40, '오프라인', 50, '=B7*D7']
##